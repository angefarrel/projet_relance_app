import openpyxl
import json
import datetime
from openpyxl.utils import get_column_letter, column_index_from_string

class PatientExtractXLSX:
    WB = None
    patients = {}
    headers = []

    # Init workbook with filename
    def __init__(self, filename):
        self.filename = filename
        self.WB = openpyxl.load_workbook(filename)

    # get selected sheet
    def get_sheet(self, name=""):
        if name:
            return self.WB.get_sheet_by_name(name)
        return self.WB.active
    
    # Get content header (each key)
    def get_header(self, pos_row_min:int = 0, pos_row_max:int=0, pos_col_min:int = 0, pos_col_max:int=0, **kwargs) -> list:
        sheet = self.get_sheet()
        for value in sheet.iter_rows(min_row=pos_row_min, max_row=pos_row_max, min_col=pos_col_min, max_col=pos_col_max, **kwargs):
            self.headers = list(value)
        return self.headers
    
    # get content for each line
    def get_file_content(self, pos_row_min:int = 0, pos_row_max:int=0, pos_col_min:int = 0, pos_col_max:int=0, **kwargs):
        for row in self.get_sheet().iter_rows(min_row=pos_row_min, max_row=pos_row_max, min_col=pos_col_min, max_col=pos_row_max, **kwargs):
            patient_code = row[0]
            patient_infos = {}

        for i in range(len(self.headers)):
            if not patient_infos.get(self.headers[i]):
                patient_infos[self.headers[i]] = row[i]
        self.patients[patient_code] = patient_infos

        # format datetime to value
        def default(obj):
            if isinstance(obj, (datetime.date, datetime.datetime)):
                return obj.isoformat()

        result = json.dumps(self.patients, indent=4, default=default)
        return result


